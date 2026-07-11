"""
Builds a human-readable Excel overview of every parsed dataset in Cerynia/DataLib,
grouped by DataLib subfolder/category (e.g. DY, DY_W, ...), then into
subsections by experiment/source file within each category.

Output: /data/arTeMiDe_Repository/Cerynia/DataLib/DataLib_Overview.xlsx

Re-run this whenever new datasets are added to DataLib/ -- it re-derives all
numeric columns (point count, Q-range, pT-range, process codes) directly from
the saved CSVs, so it never goes stale except for the hand-written columns
(experiment name, physical-process string, comments), which are looked up
from the <CATEGORY>_SECTIONS tables below and must be extended by hand for
new datasets. To add a new category (new DataLib subfolder), add a
"<CATEGORY>_SECTIONS" list following the same pattern as DY_SECTIONS /
DY_W_SECTIONS, then add it to the CATEGORIES list with its subfolder name.
"""

import sys
sys.path.append("/data/arTeMiDe_Repository/Cerynia")

from Cerynia import DataSet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DATALIB_ROOT = "/data/arTeMiDe_Repository/Cerynia/DataLib/"
OUTPUT = "/data/arTeMiDe_Repository/Cerynia/DataLib/DataLib_Overview.xlsx"

COLUMNS = [
    "Dataset", "Experiment", "Reference", "Process", "Process (artemide)",
    "Normalized", "N points", "Q min [GeV]", "Q max [GeV]",
    "pT/qT min [GeV]", "pT/qT max [GeV]", "Comments",
]

# Each category: (category_title, subfolder_under_DataLib, [section, ...])
# Each section: (title, [(dataset_name, experiment, physical_process, comment), ...])
# comment is "" unless there is a known/open issue to flag.
DY_SECTIONS = [
    ("Tevatron (CDF / D0)", [
        ("CDF1", "CDF",  "p + pbar -> Z/gamma* -> l+l-", ""),
        ("CDF2", "CDF",  "p + pbar -> Z/gamma* -> l+l-", ""),
        ("D01",  "D0",   "p + pbar -> Z/gamma* -> l+l-", ""),
        ("D02",  "D0",   "p + pbar -> Z/gamma* -> l+l-", ""),
        ("D02m", "D0",   "p + pbar -> Z/gamma* -> l+l-", ""),
    ]),
    ("ATLAS 7 TeV", [
        ("A7-00y10", "ATLAS (7 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("A7-10y20", "ATLAS (7 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("A7-20y24", "ATLAS (7 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
    ]),
    ("ATLAS 8 TeV", [
        ("A8-00y04",   "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("A8-04y08",   "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("A8-08y12",   "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("A8-12y16",   "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("A8-16y20",   "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("A8-20y24",   "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("A8-46Q66",   "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("A8-116Q150", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
    ]),
    ("ATLAS 8 TeV, normalized to 1/sigma", [
        ("A8-00y04-norm",   "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("A8-04y08-norm",   "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("A8-08y12-norm",   "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("A8-12y16-norm",   "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("A8-16y20-norm",   "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("A8-20y24-norm",   "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("A8-46Q66-norm",   "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("A8-116Q150-norm", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l-",
         "same underlying measurement as A8-* (absolute), normalized to 1/sigma instead"),
    ]),
    ("ATLAS 13 TeV, normalized to 1/sigma", [
        ("A13-norm", "ATLAS (13 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
    ]),
    ("CMS", [
        ("CMS7", "CMS (7 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("CMS8", "CMS (8 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
    ]),
    ("CMS 13 TeV, y-differential", [
        ("CMS13-00y04", "CMS (13 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("CMS13-04y08", "CMS (13 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("CMS13-08y12", "CMS (13 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("CMS13-12y16", "CMS (13 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("CMS13-16y24", "CMS (13 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("CMS13-00y04-norm", "CMS (13 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("CMS13-04y08-norm", "CMS (13 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("CMS13-08y12-norm", "CMS (13 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("CMS13-12y16-norm", "CMS (13 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("CMS13-16y24-norm", "CMS (13 TeV)", "p + p -> Z/gamma* -> l+l-",
         "pT bin edges restored from a hardcoded table (raw file only gives bin-center labels), not re-derivable from source"),
    ]),
    ("CMS 13 TeV, large Q", [
        ("CMS13_dQ_50to76",    "CMS (13 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("CMS13_dQ_76to106",   "CMS (13 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("CMS13_dQ_106to170",  "CMS (13 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("CMS13_dQ_170to350",  "CMS (13 TeV)", "p + p -> Z/gamma* -> l+l-", ""),
        ("CMS13_dQ_350to1000", "CMS (13 TeV)", "p + p -> Z/gamma* -> l+l-",
         "preliminary CMS-PAS result (hand-typed table); thFactor uses hardcoded per-point FSR correction factors, not re-derivable from source"),
    ]),
    ("LHCb", [
        ("LHCb7",  "LHCb (7 TeV)",  "p + p -> Z/gamma* -> l+l-", ""),
        ("LHCb8",  "LHCb (8 TeV)",  "p + p -> Z/gamma* -> l+l-", ""),
        ("LHCb13(2016)", "LHCb (13 TeV)", "p + p -> Z/gamma* -> l+l-",
         "hand-typed table (pre-official HepData release); provisional. Renamed from plain 'LHCb13' to disambiguate from LHCb13(2021)"),
    ]),
    ("LHCb 13 TeV, 2021 update", [
        ("LHCb13_dy", "LHCb (13 TeV)", "p + p -> Z/gamma* -> l+l-",
         "double-differential in (y,qT); '(2021)' label dropped from the name per user instruction"),
        ("LHCb13(2021)", "LHCb (13 TeV)", "p + p -> Z/gamma* -> l+l-",
         "y-integrated companion to LHCb13_dy (same paper); '(2021)' label kept per user instruction"),
    ]),
    ("RHIC (PHENIX / STAR)", [
        ("PHE200", "PHENIX (200 GeV)", "p + p -> gamma* -> l+l-",
         "invariant-xsec thFactor (qT^2 Jacobian, fixed-target-style formula) despite p+p collider process code"),
        ("STAR510", "STAR (510 GeV)", "p + p -> Z/gamma* -> l+l-",
         "qT bin edges restored from a hardcoded table (raw file only gives measured centroid), not re-derivable from source"),
    ]),
    ("E288 (fixed-target)", [
        ("E288-200", "E288", "p + Cu -> gamma* -> mu+mu- X",
         "ps_def corrected 2->1 per DataProcessor.git 'Corrections in E228' (2026-01-02). thFactor comment still says '0.001 for nb' but *1000 is applied; ported as coded, not corrected"),
        ("E288-300", "E288", "p + Cu -> gamma* -> mu+mu- X",
         "ps_def corrected 2->1 (same fix as E288-200). thFactor comment mismatch (see E288-200) still unresolved"),
        ("E288-400", "E288", "p + Cu -> gamma* -> mu+mu- X",
         "ps_def corrected 2->1 (same fix as E288-200). thFactor comment mismatch (see E288-200) still unresolved"),
    ]),
    ("E772 (fixed-target)", [
        ("E772", "E772", "p + D -> gamma* -> mu+mu- X",
         "thFactor gained the missing Feynman-x Jacobian (ffactor) per DataProcessor.git 'Corrected thFactor in E605 and E772' (2026-01-05)"),
    ]),
    ("E605 (fixed-target)", [
        ("E605", "E605", "p + Cu -> gamma* -> mu+mu- X",
         "old file mislabeled this target with proc_id=102 (the E772/deuteron code); fixed. thFactor also gained the missing Feynman-x Jacobian (ffactor, xF=0.1 fixed), same 2026-01-05 commit as E772"),
    ]),
    ("E537 (fixed-target)", [
        ("E537-dQ",  "E537", "pbar + W -> gamma* -> mu+mu- X", ""),
        ("E537-dxF", "E537", "pbar + W -> gamma* -> mu+mu- X",
         "h_1/h_2 order was swapped vs dQ in the original main-file parsing; fixed to match dQ (h_1=-1 pbar, h_2=184074 W), per DataProcessor.git's low-energy.py using that order consistently"),
    ]),
]

DY_W_SECTIONS = [
    ("Tevatron (D0 / CDF), run 1", [
        ("D01_W", "D0 (1.8 TeV)", "p + pbar -> W -> e nu",
         "Q-window is qT-dependent (kinematic bound M^2>E1+E2, per old parsing's comment), not a fixed range. Renamed from old 'D0run1-W' to match DY branch's D01 naming (same experiment, W final state)"),
        ("CDF1_W", "CDF (1.8 TeV)", "p + pbar -> W -> e nu",
         "qT bins restored from a hardcoded variable-width binning function around each point's centroid ('high-qT bins assumed from the plot'), not re-derivable from source. Rapidity restriction placed in cutParams (not y, unlike D01_W) -- inconsistent between the two datasets but ported as-is. Renamed from old 'CDFrun1-W' to match DY branch's CDF1 naming"),
    ]),
    ("CMS 8 TeV", [
        ("CMS8_W-electron", "CMS (8 TeV)", "p + p -> W -> e nu", ""),
        ("CMS8_W-muon", "CMS (8 TeV)", "p + p -> W -> mu nu",
         "cutParams' eta window (-2.5,2.5) does not match this dataset's own y-window (-2.1,2.1); ported verbatim from old parsing, not corrected"),
    ]),
]

SIDIS_SECTIONS = [
    ("HERMES 3D (z,x,pT), proton target", [
        ("hermes3D.p.pi+", "HERMES", "e + p -> e + pi+ X", ""),
        ("hermes3D.p.pi-", "HERMES", "e + p -> e + pi- X", ""),
        ("hermes3D.p.k+",  "HERMES", "e + p -> e + K+ X", ""),
        ("hermes3D.p.k-",  "HERMES", "e + p -> e + K- X", ""),
        ("hermes3D.p.pi+.no-vmsub", "HERMES", "e + p -> e + pi+ X", "no vector-meson-background subtraction (vmsub is the default/unlabeled case above)"),
        ("hermes3D.p.pi-.no-vmsub", "HERMES", "e + p -> e + pi- X", "no vector-meson-background subtraction"),
        ("hermes3D.p.k+.no-vmsub",  "HERMES", "e + p -> e + K+ X", "no vector-meson-background subtraction"),
        ("hermes3D.p.k-.no-vmsub",  "HERMES", "e + p -> e + K- X",
         "no vector-meson-background subtraction. All hermes3D sets: rows with all-zero fields (unpopulated (x,z,pT) grid cells) dropped per user confirmation; thFactor read from a raw DIS-normalization column (MSHT20nnlo), not computed from bin widths alone"),
    ]),
    ("HERMES 3D (z,x,pT), deuteron target", [
        ("hermes3D.d.pi+", "HERMES", "e + d -> e + pi+ X", ""),
        ("hermes3D.d.pi-", "HERMES", "e + d -> e + pi- X", ""),
        ("hermes3D.d.k+",  "HERMES", "e + d -> e + K+ X", ""),
        ("hermes3D.d.k-",  "HERMES", "e + d -> e + K- X", ""),
        ("hermes3D.d.pi+.no-vmsub", "HERMES", "e + d -> e + pi+ X", "no vector-meson-background subtraction"),
        ("hermes3D.d.pi-.no-vmsub", "HERMES", "e + d -> e + pi- X", "no vector-meson-background subtraction"),
        ("hermes3D.d.k+.no-vmsub",  "HERMES", "e + d -> e + K+ X", "no vector-meson-background subtraction"),
        ("hermes3D.d.k-.no-vmsub",  "HERMES", "e + d -> e + K- X", "no vector-meson-background subtraction. M_target uses the proton mass even for this deuteron-target set (per-nucleon SIDIS convention, ported verbatim)"),
    ]),
    ("COMPASS, deuteron target", [
        ("compass.d.h+", "COMPASS", "mu + d -> mu + h+ X", ""),
        ("compass.d.h-", "COMPASS", "mu + d -> mu + h- X",
         "isoscalar unidentified charged hadron; M_product uses the pion mass as a stand-in (as in old parsing)"),
    ]),
]

DY_ANGULAR_SECTIONS = [
    ("ATLAS 8 TeV, angular coefficients A0-A2", [
        ("A8-A0-00y10", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A0)", ""),
        ("A8-A0-10y20", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A0)", ""),
        ("A8-A0-20y35", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A0)", ""),
        ("A8-A1-00y10", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A1)", ""),
        ("A8-A1-10y20", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A1)", "no A1 measurement for 2<|y|<3.5"),
        ("A8-A2-00y10", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A2)", ""),
        ("A8-A2-10y20", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A2)", ""),
        ("A8-A2-20y35", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A2)", ""),
    ]),
    ("ATLAS 8 TeV, angular coefficients A3-A5", [
        ("A8-A3-00y10", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A3)", ""),
        ("A8-A3-10y20", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A3)", ""),
        ("A8-A3-20y35", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A3)", ""),
        ("A8-A4-00y10", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A4)", ""),
        ("A8-A4-10y20", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A4)", ""),
        ("A8-A4-20y35", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A4)", ""),
        ("A8-A5-00y10", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A5)", ""),
        ("A8-A5-10y20", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A5)", ""),
        ("A8-A5-20y35", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A5)", ""),
    ]),
    ("ATLAS 8 TeV, angular coefficients A6-A7", [
        ("A8-A6-00y10", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A6)", ""),
        ("A8-A6-10y20", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A6)", "no A6 measurement for 2<|y|<3.5"),
        ("A8-A7-00y10", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A7)", ""),
        ("A8-A7-10y20", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A7)", ""),
        ("A8-A7-20y35", "ATLAS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A7)",
         "ratio observables (thFactor=1); all points carry weight-process columns (ps_def_weight,h_1_weight,h_2_weight,proc_id_weight)=(1,1,1,3) so artemide can build the normalization. Old library's duplicate 'A8_Auu_*' sets (byte-identical to A4) dropped per user instruction"),
    ]),
    ("CMS 8 TeV, angular coefficients A0-A4", [
        ("CMS8-A0-00y10", "CMS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A0)", ""),
        ("CMS8-A0-10y21", "CMS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A0)", ""),
        ("CMS8-A1-00y10", "CMS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A1)", ""),
        ("CMS8-A1-10y21", "CMS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A1)", ""),
        ("CMS8-A2-00y10", "CMS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A2)", ""),
        ("CMS8-A2-10y21", "CMS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A2)", ""),
        ("CMS8-A3-00y10", "CMS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A3)", ""),
        ("CMS8-A3-10y21", "CMS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A3)", ""),
        ("CMS8-A4-00y10", "CMS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A4)", ""),
        ("CMS8-A4-10y21", "CMS (8 TeV)", "p + p -> Z/gamma* -> l+l- (A4)",
         "ratio observables; weight-process columns=(1,1,1,3), same mechanism as ATLAS8. Old library's duplicate 'CMS8-Auu-*' sets (byte-identical to A0) dropped. thFactor set to 1 per user decision, overriding old parsing's inconsistent 2/(qT_max-qT_min); no A5/A6/A7 measured in this paper"),
    ]),
]

PIDY_SECTIONS = [
    ("E537 (pion beam, fixed-target)", [
        ("E537.pi.dQ", "E537", "pi- + W -> gamma* -> mu+mu- X", ""),
        ("E537.pi.dxF", "E537", "pi- + W -> gamma* -> mu+mu- X",
         "y_min/y_max fields hold the Feynman-x (xF) window, not rapidity (ported verbatim from old parsing; no dedicated xF field in the schema). h_2 updated to the modern W nuclear code 184074 and proc_id updated to 1, replacing old parsing's [2,2,1,103], per user instruction"),
    ]),
    ("E615 (pion beam, fixed-target)", [
        ("E615.pi.dQ", "E615", "pi- + W -> gamma* -> mu+mu- X", ""),
        ("E615.pi.dxF", "E615", "pi- + W -> gamma* -> mu+mu- X",
         "y_min/y_max fields hold the Feynman-x (xF) window, not rapidity, same as E537. Old library's 20 single-bin child datasets (dQ-4.05...11.70, dxF-0.0...0.9, pure subsets of these two full datasets) not ported, per user instruction"),
    ]),
]

SIVERS_SECTIONS = [
    ("STAR (DY process, W/Z A_N)", [
        ("star23.sivers.Z", "STAR", "p + p -> Z0 -> e+e- (A_N)", ""),
        ("star26.sivers.W+.dy", "STAR", "p + p -> W+ -> e+ nu (A_N)",
         "unpublished/private-communication data (reference '????.????', as in old parsing)"),
        ("star26.sivers.W-.dy", "STAR", "p + p -> W- -> e- nu (A_N)",
         "unpublished/private-communication data. stat/syst errors fixed vs old parsing (values were shifted by one position relative to the script's own labeled source-data comment); weight-process columns (ps_def_weight,h_1_weight,h_2_weight,proc_id_weight)=(1,1,1,3/4/5 for Z/W+/W-) let artemide build the A_N ratio. Older 'star.sivers.W+/-.dqT' variant (Table4.csv-based) is preliminary and not ported"),
    ]),
    ("JLab (SIDIS process, neutron target)", [
        ("jlab.sivers.pi+", "JLab Hall A", "e + n -> e + pi+ X (A_UT Sivers)", ""),
        ("jlab.sivers.pi-", "JLab Hall A", "e + n -> e + pi- X (A_UT Sivers)", ""),
        ("jlab.sivers.k+",  "JLab Hall A", "e + n -> e + k+ X (A_UT Sivers)", ""),
        ("jlab.sivers.k-",  "JLab Hall A", "e + n -> e + k- X (A_UT Sivers)",
         "target updated to modern neutron code h_1=11 (old parsing used h_1=1/proton despite its own 'neutron target' comment and the raw files' own 'target: neutron' label); proc_id updated to 12001, weight proc_id_weight=2001 (standing SIDIS Sivers weight code), per user instruction. x/z/pT/Q bins are fixed, hand-eyeballed ranges (not real per-point edges) -- data must be evaluated at the <x>/<z>/<pT>/<Q> point, not integrated over these bins, per old parsing's own comment"),
    ]),
    ("HERMES 2009 (SIDIS process, proton target), Q-integrated", [
        ("hermes09.sivers.pi+.Qint.dz", "HERMES", "e + p -> e + pi+ X (A_UT Sivers)", ""),
        ("hermes09.sivers.pi+.Qint.dx", "HERMES", "e + p -> e + pi+ X (A_UT Sivers)", ""),
        ("hermes09.sivers.pi+.Qint.dpt", "HERMES", "e + p -> e + pi+ X (A_UT Sivers)", ""),
        ("hermes09.sivers.pi0.Qint.dz", "HERMES", "e + p -> e + pi0 X (A_UT Sivers)", ""),
        ("hermes09.sivers.pi0.Qint.dx", "HERMES", "e + p -> e + pi0 X (A_UT Sivers)", ""),
        ("hermes09.sivers.pi0.Qint.dpt", "HERMES", "e + p -> e + pi0 X (A_UT Sivers)", ""),
        ("hermes09.sivers.pi-.Qint.dz", "HERMES", "e + p -> e + pi- X (A_UT Sivers)", ""),
        ("hermes09.sivers.pi-.Qint.dx", "HERMES", "e + p -> e + pi- X (A_UT Sivers)", ""),
        ("hermes09.sivers.pi-.Qint.dpt", "HERMES", "e + p -> e + pi- X (A_UT Sivers)", ""),
        ("hermes09.sivers.k+.Qint.dz", "HERMES", "e + p -> e + k+ X (A_UT Sivers)", ""),
        ("hermes09.sivers.k+.Qint.dx", "HERMES", "e + p -> e + k+ X (A_UT Sivers)", ""),
        ("hermes09.sivers.k+.Qint.dpt", "HERMES", "e + p -> e + k+ X (A_UT Sivers)", ""),
        ("hermes09.sivers.k-.Qint.dz", "HERMES", "e + p -> e + k- X (A_UT Sivers)", ""),
        ("hermes09.sivers.k-.Qint.dx", "HERMES", "e + p -> e + k- X (A_UT Sivers)", ""),
        ("hermes09.sivers.k-.Qint.dpt", "HERMES", "e + p -> e + k- X (A_UT Sivers)",
         "renamed from old 'hermes.sivers.*' to 'hermes09.sivers.*' (2009 publication year) per user instruction, to distinguish from the newer hermes3D.sivers.* data. x/z/pT bins are fixed, hand-eyeballed ranges (not real per-point edges) -- data must be evaluated at the point, per old parsing's own comment"),
    ]),
    ("HERMES 2009 (SIDIS process, proton target), Q<2 / Q>2 sliced", [
        ("hermes09.sivers.pi+.Q<2.dz", "HERMES", "e + p -> e + pi+ X (A_UT Sivers)", ""),
        ("hermes09.sivers.pi+.Q<2.dpt", "HERMES", "e + p -> e + pi+ X (A_UT Sivers)", ""),
        ("hermes09.sivers.pi+.Q>2.dz", "HERMES", "e + p -> e + pi+ X (A_UT Sivers)",
         "Q_max fixed vs old parsing: sqrt(xMax*yMax*sM2) applied (old code omitted the sqrt, giving an implausible Q_max~19.7 GeV); confirmed independent/non-duplicate data before fixing"),
        ("hermes09.sivers.pi+.Q>2.dpt", "HERMES", "e + p -> e + pi+ X (A_UT Sivers)", "same Q_max bugfix as pi+.Q>2.dz"),
        ("hermes09.sivers.k+.Q<2.dz", "HERMES", "e + p -> e + k+ X (A_UT Sivers)", ""),
        ("hermes09.sivers.k+.Q<2.dpt", "HERMES", "e + p -> e + k+ X (A_UT Sivers)", ""),
        ("hermes09.sivers.k+.Q>2.dz", "HERMES", "e + p -> e + k+ X (A_UT Sivers)", "same Q_max bugfix as pi+.Q>2.dz"),
        ("hermes09.sivers.k+.Q>2.dpt", "HERMES", "e + p -> e + k+ X (A_UT Sivers)", "same Q_max bugfix as pi+.Q>2.dz"),
    ]),
    ("HERMES 3D (x,z,pT), proton target", [
        ("hermes3D.sivers.pi+", "HERMES", "e + p -> e + pi+ X (A_UT Sivers)", ""),
        ("hermes3D.sivers.pi-", "HERMES", "e + p -> e + pi- X (A_UT Sivers)", ""),
        ("hermes3D.sivers.k+", "HERMES", "e + p -> e + k+ X (A_UT Sivers)", ""),
        ("hermes3D.sivers.k-", "HERMES", "e + p -> e + k- X (A_UT Sivers)",
         "renamed from old 'hermes.sivers.<hadron>.3d' to 'hermes3D.sivers.<hadron>' per user instruction. Real per-point x/z/pT bin edges (unlike hermes09.*'s fixed shared bins); reference 2007.07755, distinct newer measurement from hermes09.*'s 0906.3918"),
    ]),
    ("COMPASS 2023 (SIDIS process, deuteron target)", [
        ("compass23.sivers.h+.dx", "COMPASS", "mu + d -> mu + h+ X (A_UT Sivers)", ""),
        ("compass23.sivers.h+.dz", "COMPASS", "mu + d -> mu + h+ X (A_UT Sivers)", ""),
        ("compass23.sivers.h+.dpt", "COMPASS", "mu + d -> mu + h+ X (A_UT Sivers)", ""),
        ("compass23.sivers.h-.dx", "COMPASS", "mu + d -> mu + h- X (A_UT Sivers)", ""),
        ("compass23.sivers.h-.dz", "COMPASS", "mu + d -> mu + h- X (A_UT Sivers)", ""),
        ("compass23.sivers.h-.dpt", "COMPASS", "mu + d -> mu + h- X (A_UT Sivers)",
         "target corrected to deuteron h_1=12 (old parsing used h_1=1/proton despite the raw file's own 'COMPASS deuteron 2022 data' title); proc_id updated to 12001, weight proc_id_weight=2001 (same uniform SIDIS-Sivers codes as JLab/HERMES), per user instruction. Isoscalar hadron h_2=+-12, M_product uses the pion mass as a stand-in (as in old parsing and compass.d.h+/h- in the unpolarized-SIDIS category)"),
    ]),
    ("COMPASS 2008 (SIDIS process, deuteron target), all hadrons", [
        ("compass08.sivers.pi+.dx", "COMPASS", "mu + d -> mu + pi+ X (A_UT Sivers)", ""),
        ("compass08.sivers.pi+.dz", "COMPASS", "mu + d -> mu + pi+ X (A_UT Sivers)", ""),
        ("compass08.sivers.pi+.dpt", "COMPASS", "mu + d -> mu + pi+ X (A_UT Sivers)", ""),
        ("compass08.sivers.pi-.dx", "COMPASS", "mu + d -> mu + pi- X (A_UT Sivers)", ""),
        ("compass08.sivers.pi-.dz", "COMPASS", "mu + d -> mu + pi- X (A_UT Sivers)", ""),
        ("compass08.sivers.pi-.dpt", "COMPASS", "mu + d -> mu + pi- X (A_UT Sivers)", ""),
        ("compass08.sivers.k+.dx", "COMPASS", "mu + d -> mu + k+ X (A_UT Sivers)", ""),
        ("compass08.sivers.k+.dz", "COMPASS", "mu + d -> mu + k+ X (A_UT Sivers)", ""),
        ("compass08.sivers.k+.dpt", "COMPASS", "mu + d -> mu + k+ X (A_UT Sivers)", ""),
        ("compass08.sivers.k-.dx", "COMPASS", "mu + d -> mu + k- X (A_UT Sivers)", ""),
        ("compass08.sivers.k-.dz", "COMPASS", "mu + d -> mu + k- X (A_UT Sivers)", ""),
        ("compass08.sivers.k-.dpt", "COMPASS", "mu + d -> mu + k- X (A_UT Sivers)", ""),
        ("compass08.sivers.k0.dx", "COMPASS", "mu + d -> mu + k0 X (A_UT Sivers)", ""),
        ("compass08.sivers.k0.dz", "COMPASS", "mu + d -> mu + k0 X (A_UT Sivers)", ""),
        ("compass08.sivers.k0.dpt", "COMPASS", "mu + d -> mu + k0 X (A_UT Sivers)",
         "target corrected to deuteron h_1=12 (old parsing used h_1=1/proton despite the raw file's own 'muon-deuteron DIS' title); proc_id updated to 12001, weight proc_id_weight=2001, per user instruction. h_2=20 for k0 (new code, distinct from the old scheme's h_2=3). Collins-asymmetry siblings from the same old script not ported (Sivers only)"),
    ]),
    ("COMPASS 2008 (SIDIS process, deuteron target), leading hadrons", [
        ("compass08.sivers.pi+.dx.leading", "COMPASS", "mu + d -> mu + pi+ X (A_UT Sivers, leading)", ""),
        ("compass08.sivers.pi+.dz.leading", "COMPASS", "mu + d -> mu + pi+ X (A_UT Sivers, leading)", ""),
        ("compass08.sivers.pi+.dpt.leading", "COMPASS", "mu + d -> mu + pi+ X (A_UT Sivers, leading)", ""),
        ("compass08.sivers.pi-.dx.leading", "COMPASS", "mu + d -> mu + pi- X (A_UT Sivers, leading)", ""),
        ("compass08.sivers.pi-.dz.leading", "COMPASS", "mu + d -> mu + pi- X (A_UT Sivers, leading)", ""),
        ("compass08.sivers.pi-.dpt.leading", "COMPASS", "mu + d -> mu + pi- X (A_UT Sivers, leading)", ""),
        ("compass08.sivers.k+.dx.leading", "COMPASS", "mu + d -> mu + k+ X (A_UT Sivers, leading)", ""),
        ("compass08.sivers.k+.dz.leading", "COMPASS", "mu + d -> mu + k+ X (A_UT Sivers, leading)", ""),
        ("compass08.sivers.k+.dpt.leading", "COMPASS", "mu + d -> mu + k+ X (A_UT Sivers, leading)", ""),
        ("compass08.sivers.k-.dx.leading", "COMPASS", "mu + d -> mu + k- X (A_UT Sivers, leading)", ""),
        ("compass08.sivers.k-.dz.leading", "COMPASS", "mu + d -> mu + k- X (A_UT Sivers, leading)", ""),
        ("compass08.sivers.k-.dpt.leading", "COMPASS", "mu + d -> mu + k- X (A_UT Sivers, leading)", ""),
        ("compass08.sivers.k0.dx.leading", "COMPASS", "mu + d -> mu + k0 X (A_UT Sivers, leading)", ""),
        ("compass08.sivers.k0.dz.leading", "COMPASS", "mu + d -> mu + k0 X (A_UT Sivers, leading)", ""),
        ("compass08.sivers.k0.dpt.leading", "COMPASS", "mu + d -> mu + k0 X (A_UT Sivers, leading)",
         "leading-hadron sample: old on-disk library never included these (only 'all hadrons' sets existed), even though the old script builds them with active SaveToCSV calls -- ported per user instruction, using the same corrected process code as the main sets. Different z_min cut (0.25 vs 0.2) than the 'all hadrons' sets, not a duplicate. Naming: old 'compass08.sivers.pi+leading.dx' -> 'compass08.sivers.pi+.dx.leading' (.leading moved to the end), per user instruction"),
    ]),
    ("COMPASS 2016 (SIDIS process, proton target), joined across Q-windows", [
        ("compass16.sivers.h-.1<z<2.dpt", "COMPASS", "mu + p -> mu + h- X (A_UT Sivers)", ""),
        ("compass16.sivers.h-.1<z<2.dx", "COMPASS", "mu + p -> mu + h- X (A_UT Sivers)", ""),
        ("compass16.sivers.h-.1<z<2.dz", "COMPASS", "mu + p -> mu + h- X (A_UT Sivers)", ""),
        ("compass16.sivers.h-.1<z.dpt", "COMPASS", "mu + p -> mu + h- X (A_UT Sivers)", ""),
        ("compass16.sivers.h-.1<z.dx", "COMPASS", "mu + p -> mu + h- X (A_UT Sivers)", ""),
        ("compass16.sivers.h-.1<z.dz", "COMPASS", "mu + p -> mu + h- X (A_UT Sivers)", ""),
        ("compass16.sivers.h-.2<z.dpt", "COMPASS", "mu + p -> mu + h- X (A_UT Sivers)", ""),
        ("compass16.sivers.h-.2<z.dx", "COMPASS", "mu + p -> mu + h- X (A_UT Sivers)", ""),
        ("compass16.sivers.h-.2<z.dz", "COMPASS", "mu + p -> mu + h- X (A_UT Sivers)", ""),
        ("compass16.sivers.h+.1<z<2.dpt", "COMPASS", "mu + p -> mu + h+ X (A_UT Sivers)", ""),
        ("compass16.sivers.h+.1<z<2.dx", "COMPASS", "mu + p -> mu + h+ X (A_UT Sivers)", ""),
        ("compass16.sivers.h+.1<z<2.dz", "COMPASS", "mu + p -> mu + h+ X (A_UT Sivers)", ""),
        ("compass16.sivers.h+.1<z.dpt", "COMPASS", "mu + p -> mu + h+ X (A_UT Sivers)", ""),
        ("compass16.sivers.h+.1<z.dx", "COMPASS", "mu + p -> mu + h+ X (A_UT Sivers)", ""),
        ("compass16.sivers.h+.1<z.dz", "COMPASS", "mu + p -> mu + h+ X (A_UT Sivers)", ""),
        ("compass16.sivers.h+.2<z.dpt", "COMPASS", "mu + p -> mu + h+ X (A_UT Sivers)", ""),
        ("compass16.sivers.h+.2<z.dx", "COMPASS", "mu + p -> mu + h+ X (A_UT Sivers)", ""),
        ("compass16.sivers.h+.2<z.dz", "COMPASS", "mu + p -> mu + h+ X (A_UT Sivers)",
         "each dataset is double-differential in Q (4 windows: 1-2, 2-2.5, 2.5-4, 4-9 GeV) and the named variable -- joined from what were 4 separate 'binQ1'-'binQ4' files in the old on-disk library (old ReadSSA_Compass16_v2.py already built this joined form internally; only renamed here with the .dQ suffix, per user instruction, replacing the redundant per-Q-bin child datasets). Target is PROTON (h_1=1) -- same COMPASS 2010 proton measurement as wgt/compass16.wgt.* (same raw data folder, same paper PLB770(2017)138/1609.07374, confirmed via the paper's own abstract after an initial wrong deuteron guess by analogy with compass08/23.sivers.*). proc_id=12001/weight=2001. A third, oldest script (ReadSSA_Compass16.py) was found to be a byte-identical-data precursor with an incomplete process code -- not used at all"),
    ]),
]

WGT_SECTIONS = [
    ("COMPASS 2016 pub. (SIDIS process, proton target, 2010 data), joined across Q-windows", [
        ("compass16.wgt.h-.1<z<2.dpt", "COMPASS", "mu + p -> mu + h- X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h-.1<z<2.dx", "COMPASS", "mu + p -> mu + h- X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h-.1<z<2.dz", "COMPASS", "mu + p -> mu + h- X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h-.1<z.dpt", "COMPASS", "mu + p -> mu + h- X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h-.1<z.dx", "COMPASS", "mu + p -> mu + h- X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h-.1<z.dz", "COMPASS", "mu + p -> mu + h- X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h-.2<z.dpt", "COMPASS", "mu + p -> mu + h- X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h-.2<z.dx", "COMPASS", "mu + p -> mu + h- X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h-.2<z.dz", "COMPASS", "mu + p -> mu + h- X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h+.1<z<2.dpt", "COMPASS", "mu + p -> mu + h+ X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h+.1<z<2.dx", "COMPASS", "mu + p -> mu + h+ X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h+.1<z<2.dz", "COMPASS", "mu + p -> mu + h+ X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h+.1<z.dpt", "COMPASS", "mu + p -> mu + h+ X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h+.1<z.dx", "COMPASS", "mu + p -> mu + h+ X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h+.1<z.dz", "COMPASS", "mu + p -> mu + h+ X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h+.2<z.dpt", "COMPASS", "mu + p -> mu + h+ X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h+.2<z.dx", "COMPASS", "mu + p -> mu + h+ X (A_LT^cos(phi_h-phi_S))", ""),
        ("compass16.wgt.h+.2<z.dz", "COMPASS", "mu + p -> mu + h+ X (A_LT^cos(phi_h-phi_S))",
         "target is PROTON (raw files' own header says 'COMPASS proton 2010 data') -- same COMPASS 2010 proton measurement as Sivers/compass16.sivers.* (same raw data folder /COMPASS/1609.07374/, same paper; the '16' in the name is the publication year, not the data-collection year). proc_id=13001 (new, distinguishing A_LT from Sivers' 12001), weight proc_id=2001 (same weighting process as Sivers). Renamed from old 'compass16.ALT.*' to 'compass16.wgt.*' (matching this category's folder name) and joined across the 4 Q-windows with a .dQ suffix, same treatment as Sivers/COMPASS16.py. normErr=[0.03] is a dilution factor, ported verbatim from old parsing"),
    ]),
    ("HERMES 3D (x,z,pT), proton target", [
        ("hermes3D.wgt.pi+", "HERMES", "e + p -> e + pi+ X (A_LT^cos(phi-phis))", ""),
        ("hermes3D.wgt.pi-", "HERMES", "e + p -> e + pi- X (A_LT^cos(phi-phis))", ""),
        ("hermes3D.wgt.k+", "HERMES", "e + p -> e + k+ X (A_LT^cos(phi-phis))", ""),
        ("hermes3D.wgt.k-", "HERMES", "e + p -> e + k- X (A_LT^cos(phi-phis))",
         "h_2 fixed to -2 (old parsing had -1, a copy-paste bug from the pi- block, giving k- the same hadron code as pi-), per user instruction. Renamed from old 'hermes3D.ALT.<hadron>' to 'hermes3D.wgt.<hadron>' for naming consistency with compass16.wgt.*. Reads the same raw SFA 3D files as Sivers/hermes3D.sivers.*, a different asymmetry block within each file; proc_id=13001/weight=2001, h_1=1 (proton, already correct in old parsing, no target-species issue here)"),
    ]),
    ("JLab (6 GeV, SIDIS process, neutron target)", [
        ("JLab6.wgt.pi+", "JLab Hall A", "e + n -> e + pi+ X (A_LT)", ""),
        ("JLab6.wgt.pi-", "JLab Hall A", "e + n -> e + pi- X (A_LT)",
         "target updated to neutron h_1=11 (old parsing used h_1=1/proton despite its own 'off neutron' comment, same fix as Sivers/JLab.py); proc_id updated to 13001, weight proc_id=2001 (uniform wgt codes, replacing old parsing's non-standard 13003/2003), per user instruction. Hardcoded literal table from arXiv:1108.0489 (no raw data file); Q/z/pT bins are fixed literals from the paper text, x bins are hand-computed fixed ranges -- not real per-point edges"),
    ]),
]

G2_SECTIONS = [
    ("E142 / E143 (transverse spin structure function g2)", [
        ("E142.n", "E142", "e + n -> e + X (g2)", ""),
        ("E143.p", "E143", "e + p -> e + X (g2-bar)", ""),
        ("E143.d", "E143", "e + d -> e + X (g2-bar)", ""),
        ("E143.n", "E143", "e + n -> e + X (g2-bar)", ""),
        ("E143-1995.p", "E143", "e + p -> e + X (g2)", ""),
        ("E143-1995.d", "E143", "e + d -> e + X (g2)",
         "process code carried over unchanged from the old script's single 'process' integer (100=p,101=n,102=d), mapped to ps_def=1,h_1=1,proc_id=<old value> per user instruction -- G2 has no h_2 (inclusive DIS). This whole category is a pure reformatting port: the user confirmed the old script's physics/process definitions are already the modern standard, nothing was changed. E143.p/d/n present g2-bar directly (WW/leading-twist already removed by the experiment); E143-1995.p/d need the WW-term subtraction like the rest of this category"),
    ]),
    ("E154 / E155 (transverse spin structure function g2)", [
        ("E154.n", "E154", "e + n -> e + X (g2)",
         "reference/comment are a likely copy-paste leftover from the neighboring E155-1999 block in the old script (says 'tables 1,2,3'/'No syst. uncert.' but E154 reads 2 tables and does include a syst term) -- ported verbatim, not corrected, per the 'nothing to change' instruction"),
        ("E155-29.p", "E155 (29.1 GeV)", "e + p -> e + X (g2)", ""),
        ("E155-29.d", "E155 (29.1 GeV)", "e + d -> e + X (g2)", ""),
        ("E155-32.p", "E155 (32.3 GeV)", "e + p -> e + X (g2)", ""),
        ("E155-32.d", "E155 (32.3 GeV)", "e + d -> e + X (g2)",
         "old comment says 'tables 1,2,3' (copy-paste leftover from the 29.1 GeV deuteron block; actually reads Table4/5/6) -- ported verbatim"),
        ("E155-38.p", "E155 (38.8 GeV)", "e + p -> e + X (g2)", ""),
        ("E155-38.d", "E155 (38.8 GeV)", "e + d -> e + X (g2)",
         "thFactor=x for both E155-38 sets (unlike E155-29/32 where the deuteron sibling typically has thFactor=1 -- E155-29.d is the one exception, thFactor=x there too, verified directly against the old script's line numbers after an initial mis-extraction). E155-32 reuses E155-29's WW-correction terms, per the old script's own comment (same x/Q^2 points at both beam energies)"),
    ]),
    ("Hall A / HERMES / RSS / SMC (transverse spin structure function g2)", [
        ("HallA-2004.n", "JLab Hall A (2004)", "e + n -> e + X (g2)", ""),
        ("HallA-2016-4.He3", "JLab Hall A (2016, E=4.74 GeV)", "e + He3 -> e + X (g2)",
         "proc_id=101 (neutron) despite the He3-labeled dataset name -- per-nucleon effective-neutron extraction, old script's own inline comment says 'it is neutron!'"),
        ("HallA-2016-5.He3", "JLab Hall A (2016, E=5.89 GeV)", "e + He3 -> e + X (g2)", "same neutron-code convention as HallA-2016-4.He3"),
        ("HERMES", "HERMES", "e + p -> e + X (g2)", ""),
        ("HERMES.av", "HERMES", "e + p -> e + X (g2)",
         "Q^2-evolved/x-averaged companion to HERMES (Table 2 vs Table 1 in the paper). Both HERMES sets have a real per-point Q-bin guess computed in the old script but never actually used -- a fixed Q window is used instead (dead code, ported as found)"),
        ("RSS.p", "RSS", "e + p -> e + X (g2)", "hand-digitized from a figure (no HEPData table); fixed Q window for all points"),
        ("SMC.p", "SMC", "mu + p -> mu + X (g2)",
         "the only G2 dataset that does not need a separate WW-term subtraction -- SMC publishes both raw g2 and a WW-corrected g2 table directly, xSec is their difference"),
    ]),
]

D2_SECTIONS = [
    ("Lattice QCD (RQCD / QCDSF / GHMP26)", [
        ("RQCD_d2_ud", "RQCD (lattice)", "u, d quark d2 moment",
         "hand-typed from table 4 for u/d quarks; proc_id=2/1 (no inline u/d label in the old script, preserved as-is)"),
        ("RQCD_d2_singlet", "RQCD (lattice)", "u-d, u+d quark d2 moment",
         "proc_id=11 (u-d) / 12 (u+d), per the old script's own inline comments"),
        ("RQCD_d2_pn", "RQCD (lattice)", "p, n d2 moment (lattice)",
         "old dataset comment: 'better not to use because there are no sea-part'"),
        ("QCDSF_d2", "QCDSF (lattice)", "p, n d2 moment (lattice)", ""),
        ("GHMP26_d2", "GHMP26 (lattice)", "u-d quark d2 moment",
         "reference arXiv:2604.00143 is a forward-dated id in the old script, ported verbatim"),
    ]),
    ("E143 / E154 / E155 (transverse spin structure function d2)", [
        ("E143-1995_d2", "E143 (1995)", "e + p,d -> e + X (d2)", "2 points: p, d"),
        ("E143_d2", "E143", "e + p,d,n -> e + X (d2)", "3 points: p, d, n"),
        ("E154_d2", "E154", "e + n -> e + X (d2)",
         "s computed with M_proton despite the neutron target, ported verbatim from the old script"),
        ("E155-1999_d2", "E155 (1999)", "e + p,d -> e + X (d2)", "2 points: p, d"),
        ("E155_d2", "E155 (2002)", "e + p,d -> e + X (d2)", "2 points: p, d"),
    ]),
    ("Hall A / HERMES / RSS / SANE (transverse spin structure function d2)", [
        ("HallA-2004_d2", "JLab Hall A (2004)", "e + n -> e + X (d2)", ""),
        ("HallA-2014_d2", "JLab Hall A (2014)", "e + n -> e + X (d2)", "2 Q-bins"),
        ("HallA-2016_d2", "JLab Hall A (2016)", "e + n -> e + X (d2)",
         "same xSec/Q values as HallA-2014 with an added 3rd systematic-error column -- kept as a separate dataset, per the old script"),
        ("HERMES_d2", "HERMES", "e + p -> e + X (d2)", ""),
        ("RSS-2006_d2", "RSS (2006)", "e + p -> e + X (d2)", ""),
        ("RSS-2008_d2", "RSS (2010)", "e + p,d,n -> e + X (d2)", "3 points: p, d, n"),
        ("SANE_d2", "SANE", "e + p -> e + X (d2)",
         "thFactor=-1 (paper's own sign convention, see their eqn.(2)); each point has its own beam energy and Q window, unlike every other D2 dataset"),
    ]),
]

CATEGORIES = [
    ("DY -- unpolarized Drell-Yan", "DY", DY_SECTIONS),
    ("DY_W -- W-boson production", "DY_W", DY_W_SECTIONS),
    ("SIDIS -- unpolarized SIDIS", "SIDIS", SIDIS_SECTIONS),
    ("DY_angular -- Drell-Yan angular coefficients", "DY_angular", DY_ANGULAR_SECTIONS),
    ("piDY -- pion-beam Drell-Yan", "piDY", PIDY_SECTIONS),
    ("Sivers -- transverse single-spin asymmetries", "Sivers", SIVERS_SECTIONS),
    ("wgt -- weighted (double-spin) asymmetries", "wgt", WGT_SECTIONS),
    ("G2 -- transverse spin structure function g2", "G2", G2_SECTIONS),
    ("D2 -- transverse spin structure function d2 (moment)", "D2", D2_SECTIONS),
]

# -- Build workbook ----------------------------------------------------------

wb = Workbook()
ws = wb.active
ws.title = "DataLib overview"

title_font     = Font(bold=True, size=14)
category_font  = Font(bold=True, size=13, color="FFFFFF")
category_fill  = PatternFill("solid", fgColor="203864")
section_font   = Font(bold=True, size=11, color="FFFFFF")
section_fill   = PatternFill("solid", fgColor="4472C4")
header_font    = Font(bold=True)
header_fill    = PatternFill("solid", fgColor="D9E1F2")
thin           = Side(style="thin", color="BFBFBF")
border         = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap           = Alignment(wrap_text=True, vertical="top")
center         = Alignment(horizontal="center", vertical="center")

ncols = len(COLUMNS)

ws.cell(row=1, column=1, value="Cerynia DataLib -- parsed datasets overview").font = title_font
row = 3

for category_title, subfolder, sections in CATEGORIES:
    # Category banner
    ws.cell(row=row, column=1, value=category_title)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1)
    cell.font = category_font
    cell.fill = category_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    row += 2

    datalib_dir = DATALIB_ROOT + subfolder + "/"

    for section_title, entries in sections:
        # Section banner
        ws.cell(row=row, column=1, value=section_title)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row=row, column=1)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        # Header row
        for c, name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row, column=c, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center
        row += 1

        for dsname, experiment, process, comment in entries:
            ds = DataSet.from_csv(datalib_dir + dsname + ".csv")
            df = ds.df
            ### G2/D2 have no h_2 (inclusive DIS, no produced hadron)
            proc_cols = ["ps_def", "h_1", "h_2", "proc_id"] if "h_2" in df.columns else ["ps_def", "h_1", "proc_id"]
            procs = df[proc_cols].drop_duplicates().values.tolist()
            proc_str = "; ".join(str(p) for p in procs)
            ### weight-process columns (only present for ratio observables, e.g. DY_angular)
            weight_cols = ["ps_def_weight", "h_1_weight", "h_2_weight", "proc_id_weight"]
            if all(c in df.columns for c in weight_cols):
                wprocs = df[weight_cols].drop_duplicates().values.tolist()
                proc_str += "  |  weight: " + "; ".join(str(p) for p in wprocs)
            ### DY branch uses qT_min/qT_max; SIDIS branch uses pT_min/pT_max; G2 has neither
            ### (use x_min/x_max instead); D2 has none of the three (no x-bin at all -- it's
            ### an x-integrated moment) so pt_min_col/pt_max_col stay None and the columns are left blank
            if "qT_min" in df.columns:
                pt_min_col, pt_max_col = "qT_min", "qT_max"
            elif "pT_min" in df.columns:
                pt_min_col, pt_max_col = "pT_min", "pT_max"
            elif "x_min" in df.columns:
                pt_min_col, pt_max_col = "x_min", "x_max"
            else:
                pt_min_col, pt_max_col = None, None

            values = [
                dsname,
                experiment,
                ds.reference,
                process,
                proc_str,
                "Yes" if ds.isNormalized else "",
                len(df),
                round(float(df["Q_min"].min()), 4),
                round(float(df["Q_max"].max()), 4),
                round(float(df[pt_min_col].min()), 4) if pt_min_col else "",
                round(float(df[pt_max_col].max()), 4) if pt_max_col else "",
                comment,
            ]
            for c, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.border = border
                if c in (7, 8, 9, 10, 11):
                    cell.alignment = center
                elif c == 12:
                    cell.alignment = wrap
            row += 1

        row += 1  # blank separator row between sections

    row += 1  # extra blank row between categories

# -- Column widths, freeze header ------------------------------------------

widths = [14, 16, 26, 30, 20, 11, 9, 11, 11, 11, 11, 55]
for c, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A4"

wb.save(OUTPUT)
print(f"Wrote {OUTPUT}")
